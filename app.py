import streamlit as st
import pandas as pd
import zipfile
import io
import os
from openpyxl import load_workbook
import tempfile
from typing import Dict, List, Tuple, Optional

def extract_student_id(file_path: str) -> Optional[str]:
    """
    Extract student ID from 'Current Semester Advising' sheet, cell C5
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Student ID as string or None if not found/error
    """
    try:
        workbook = load_workbook(file_path, data_only=True)
        
        # Check if 'Current Semester Advising' sheet exists
        if 'Current Semester Advising' not in workbook.sheetnames:
            return None
            
        sheet = workbook['Current Semester Advising']
        student_id = sheet['C5'].value
        
        if student_id is not None:
            return str(student_id).strip()
        return None
        
    except Exception as e:
        st.error(f"Error extracting student ID: {str(e)}")
        return None

def extract_internship_data(file_path: str) -> Optional[Dict[str, int]]:
    """
    Extract internship data (completed hours by internship code) from Excel file
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Dictionary mapping internship codes to completed hours, or None if error
    """
    try:
        # Try to read all sheets and find the one with internship data
        excel_file = pd.ExcelFile(file_path)
        
        internship_data = {}
        
        for sheet_name in excel_file.sheet_names:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                
                # Look for the header pattern: Internship Code, Total Hours, Completed, Remaining
                for i in range(len(df)):
                    row = df.iloc[i]
                    if (len(row) >= 4 and 
                        pd.notna(row.iloc[0]) and 
                        pd.notna(row.iloc[2]) and
                        str(row.iloc[0]).strip().lower() == 'internship code' and
                        str(row.iloc[2]).strip().lower() == 'completed'):
                        
                        # Found the header row, now extract data
                        for j in range(i + 1, len(df)):
                            data_row = df.iloc[j]
                            if (len(data_row) >= 4 and 
                                pd.notna(data_row.iloc[0]) and 
                                pd.notna(data_row.iloc[2])):
                                
                                internship_code = str(data_row.iloc[0]).strip()
                                completed_hours = data_row.iloc[2]
                                
                                # Skip if internship code is empty or completed hours is not numeric
                                if internship_code and pd.notna(completed_hours):
                                    try:
                                        completed_hours = int(float(completed_hours))
                                        internship_data[internship_code] = completed_hours
                                    except (ValueError, TypeError):
                                        continue
                            else:
                                # Stop when we hit empty rows
                                break
                        
                        # Found data in this sheet, return it
                        if internship_data:
                            return internship_data
                            
            except Exception:
                continue  # Try next sheet
                
        return internship_data if internship_data else None
        
    except Exception as e:
        st.error(f"Error extracting internship data: {str(e)}")
        return None

def process_zip_file(uploaded_file) -> Tuple[pd.DataFrame, List[str], List[str]]:
    """
    Process uploaded zip file and extract student internship data
    
    Args:
        uploaded_file: Streamlit uploaded file object
        
    Returns:
        Tuple of (consolidated_dataframe, processed_files, error_files)
    """
    processed_files = []
    error_files = []
    all_student_data = []
    
    with tempfile.TemporaryDirectory() as temp_dir:
        # Extract zip file
        with zipfile.ZipFile(uploaded_file, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
            
            # Get all Excel files from the extracted directory
            excel_files = []
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    if file.endswith(('.xlsx', '.xls')):
                        excel_files.append(os.path.join(root, file))
            
            if not excel_files:
                st.error("No Excel files found in the uploaded zip file.")
                return pd.DataFrame(), [], []
            
            # Process each Excel file
            for file_path in excel_files:
                file_name = os.path.basename(file_path)
                
                try:
                    # Extract student ID
                    student_id = extract_student_id(file_path)
                    if not student_id:
                        error_files.append(f"{file_name}: Could not extract student ID from 'Current Semester Advising' sheet, cell C5")
                        continue
                    
                    # Extract internship data
                    internship_data = extract_internship_data(file_path)
                    if not internship_data:
                        error_files.append(f"{file_name}: Could not extract internship data")
                        continue
                    
                    # Add student data to consolidated list
                    student_record = {'Student_ID': student_id}
                    student_record.update(internship_data)
                    all_student_data.append(student_record)
                    
                    processed_files.append(file_name)
                    
                except Exception as e:
                    error_files.append(f"{file_name}: {str(e)}")
    
    # Create consolidated DataFrame
    if all_student_data:
        consolidated_df = pd.DataFrame(all_student_data)
        # Fill NaN values with 0 for internship codes not present in all files
        consolidated_df = consolidated_df.fillna(0)
        # Ensure Student_ID is the first column
        cols = ['Student_ID'] + [col for col in consolidated_df.columns if col != 'Student_ID']
        consolidated_df = consolidated_df[cols]
    else:
        consolidated_df = pd.DataFrame()
    
    return consolidated_df, processed_files, error_files

def main():
    st.title("Student Internship Data Consolidator")
    st.markdown("""
    Upload a zip file containing Excel files with student internship data. 
    The application will extract student IDs and consolidate completed internship hours into a single report.
    """)
    
    # File upload section
    st.header("Upload Zip File")
    uploaded_file = st.file_uploader(
        "Choose a zip file containing student Excel files",
        type=['zip'],
        help="Upload a zip file containing Excel files (.xlsx) with student internship data"
    )
    
    if uploaded_file is not None:
        st.success(f"Uploaded: {uploaded_file.name}")
        
        # Process button
        if st.button("Process Files", type="primary"):
            with st.spinner("Processing files... This may take a few moments."):
                # Process the zip file
                consolidated_df, processed_files, error_files = process_zip_file(uploaded_file)
                
                # Display results
                st.header("Processing Results")
                
                # Summary statistics
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Files Processed Successfully", len(processed_files))
                with col2:
                    st.metric("Files with Errors", len(error_files))
                with col3:
                    st.metric("Total Students", len(consolidated_df) if not consolidated_df.empty else 0)
                
                # Display errors if any
                if error_files:
                    st.subheader("⚠️ Files with Errors")
                    with st.expander("View Error Details"):
                        for error in error_files:
                            st.error(error)
                
                # Display processed files
                if processed_files:
                    st.subheader("✅ Successfully Processed Files")
                    with st.expander("View Processed Files"):
                        for file in processed_files:
                            st.success(file)
                
                # Display consolidated data
                if not consolidated_df.empty:
                    st.header("Consolidated Data Preview")
                    st.dataframe(
                        consolidated_df,
                        use_container_width=True,
                        hide_index=True
                    )
                    
                    # Download section
                    st.header("Download Consolidated Report")
                    
                    # Convert DataFrame to Excel bytes
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        consolidated_df.to_excel(writer, sheet_name='Consolidated_Report', index=False)
                    
                    excel_data = output.getvalue()
                    
                    st.download_button(
                        label="📥 Download Consolidated Excel Report",
                        data=excel_data,
                        file_name="consolidated_internship_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    # Display summary information
                    st.subheader("Report Summary")
                    st.info(f"""
                    - **Total Students**: {len(consolidated_df)}
                    - **Internship Codes Found**: {len([col for col in consolidated_df.columns if col != 'Student_ID'])}
                    - **Files Successfully Processed**: {len(processed_files)}
                    - **Files with Errors**: {len(error_files)}
                    """)
                else:
                    st.error("No data could be extracted from the uploaded files. Please check the file format and ensure they contain the required sheets and data structure.")
    
    # Instructions section
    with st.expander("📋 Instructions and Requirements"):
        st.markdown("""
        ### File Requirements:
        - Upload a **zip file** containing Excel files (.xlsx format)
        - Each Excel file should represent one student's data
        - Each file must have a sheet named **"Current Semester Advising"** with the student ID in cell **C5**
        - Each file should contain internship data with columns: **Internship Code**, **Total Hours**, **Completed**, **Remaining**
        
        ### Expected Data Format:
        The application looks for internship data in the following format:
        ```
        Internship Code | Total Hours | Completed | Remaining
        SPTH290        |     50      |     25    |    25
        SPTH291        |     50      |     30    |    20
        ...
        ```
        
        ### Output:
        - Consolidated Excel file with student IDs as rows and internship codes as columns
        - Values represent completed hours for each internship per student
        - Missing internship codes for a student will show as 0
        """)

if __name__ == "__main__":
    main()
